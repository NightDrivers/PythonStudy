# -*- coding: utf-8 -*-
import openpyxl
import os
import argparse


def parse_argv():
    __parse = argparse.ArgumentParser()
    __parse.add_argument("excel", help="注册设备Excel文件路径")
    __parse.add_argument("table", help="注册设备表名")
    __parse.add_argument("export_txt", help="导出的txt路径")
    __parse.add_argument("--device-prefix", help="添加设备名前缀", default="")
    return __parse.parse_args()


if __name__ == '__main__':
    argv = parse_argv()
    excel = argv.excel
    table = argv.table
    device_prefix = argv.device_prefix
    export_txt = argv.export_txt

    print("仅支持工作表第一列为设备名称，第二列为设备UUID的Excel文件")
    print("工作表第一行被当成标题头忽略")

    if not os.path.exists(excel):
        print('设备文件不存在')
        exit(-1)
    wb = openpyxl.load_workbook(filename=excel)
    if not wb.sheetnames.__contains__(table):
        print('未找到对应表格')
        exit(-1)
    # ws = wb.create_sheet()
    ws = wb[table]
    max_row = ws.max_row
    max_column = ws.max_column

    if max_row < 2 or max_column < 2:
        print("设备信息内容不足")
        exit(-1)

    txt_file = open(export_txt, "wt")
    txt_file.write("Device ID\tDevice Name\tDevice Platform\n")
    for i in range(2, max_row + 1):
        device_name = ws.cell(row=i, column=1).value
        if device_name is None:
            continue
        if device_prefix != "":
            device_name = "{0} {1}".format(device_prefix, device_name)
        uuid = ws.cell(row=i, column=2).value
        txt_file.write("{1}\t{0}\tios\n".format(device_name, uuid))
    txt_file.close()
