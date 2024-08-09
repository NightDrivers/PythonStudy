# -*- coding: utf-8 -*-
import openpyxl
import os
import subprocess
import argparse
from openpyxl.styles import Alignment
import re

line_localize_rex = '^".?" = ".?";\\n$|.? = ".?";\\n$'
excel_path = ""
format_string_placeholder_rex = '%@|%[#\\-+0]?[0-9*]*\\.?[0-9*]*[diouxXcsb]|%[#\\-+0]?[0-9*]*\\.?[0-9*]*L?[fFeEgGaA]'
key_discard_placeholders = ["", "不要改我"]


# 加载翻译文件内容
def load_localize_content(strings_file_path: str):
    content = dict()
    with open(strings_file_path) as file:
        for line in file:
            pairs = line.split(' = ')
            if len(pairs) == 2:
                if pairs[0].startswith('"'):
                    key = pairs[0][1:-1]
                else:
                    key = pairs[0]
                content[key] = pairs[1][1:-3]
    return content


def localize_export(root_path: str, mode_name: str, strings_name: str, sortByEmptyContent: bool):
    global excel_path
    dirs = []
    for item in os.listdir(root_path):
        if item.endswith('.lproj'):
            temp = item[:-6]
            if temp != 'Base':
                dirs.append(temp)
    if len(dirs) == 0:
        print('没有找到翻译相关文件')
        exit(-1)
    print(dirs)
    dic = dict()
    keys = set()
    keys_weight_dic = dict()
    for lang in dirs:
        path = '{0}/{1}.lproj/{2}.strings'.format(root_path, lang, strings_name)
        content_dic = load_localize_content(path)
        print("翻译文本数量 语言 {0}: {1}".format(lang, len(content_dic)))
        if sortByEmptyContent:
            for key in content_dic.keys():
                value = content_dic[key]
                if not keys_weight_dic.__contains__(key):
                    if len(value) == 0:
                        keys_weight_dic[key] = 1
                    else:
                        keys_weight_dic[key] = 0
                else:
                    weight = keys_weight_dic[key]
                    if len(value) == 0:
                        keys_weight_dic[key] = weight + 1
        keys = keys.union(set(content_dic.keys()))
        dic[lang] = content_dic
    # for item in keys:
    #     print(item)
    print("总的翻译文本数量: {0}".format(len(keys)))
    if sortByEmptyContent:
        keys_sorted = sorted(keys, key=lambda e: keys_weight_dic[e], reverse=True)
    else:
        keys_sorted = sorted(keys)

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(filename=excel_path)
        if wb.sheetnames.__contains__(mode_name):
            ws = wb[mode_name]
            wb.remove(ws)
        ws = wb.create_sheet(mode_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = mode_name
    ws.sheet_format.defaultColWidth = 50

    langs_sorted = sorted(dirs)
    try:
        langs_sorted.remove("zh-Hans")
        langs_sorted.insert(0, "zh-Hans")
    except Exception as e:
        print("zh-Hans not exist {e}")
    langs_sorted.insert(0, 'key')

    # 将翻译内容写入excel
    for i in range(0, len(langs_sorted)):
        for j in range(1, len(keys_sorted) + 2):
            temp_lang = langs_sorted[i]
            cell = ws.cell(row=j, column=i + 1)
            if j == 1:
                cell.value = temp_lang
            else:
                key = keys_sorted[j - 2]
                if i == 0:
                    cell.value = key
                else:
                    temp_dict = dic[temp_lang]
                    if temp_dict.keys().__contains__(key):
                        cell.value = temp_dict[key]
                    elif key.startswith('"'):
                        cell.value = ""
                    else:
                        # InfoPlist翻译，缺少某个key，做个标记，提醒翻译人员不要修改此项
                        # 同时作为导入时不需要翻译的标记
                        cell.value = "不要改我"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    wb.save(excel_path)


def localize_import(root_path: str, mode_name: str, strings_name: str, not_wrap_key: bool):
    global excel_path, format_string_placeholder_rex
    if not os.path.exists(excel_path):
        print('翻译文件不存在: {0}'.format(excel_path))
        exit(-1)
    wb = openpyxl.load_workbook(filename=excel_path)
    if not wb.sheetnames.__contains__(mode_name):
        print('翻译文件未找到模块: {0}'.format(mode_name))
        exit(-1)
    ws = wb[mode_name]
    print('max row {0}'.format(ws.max_row))
    print('max column {0}'.format(ws.max_column))
    max_row = ws.max_row
    max_column = ws.max_column

    for i in range(2, max_column + 1):
        lang = ws.cell(row=1, column=i).value
        if lang is None or len(lang) == 0:
            continue
        string_file_path = '{0}/{1}.lproj/{2}.strings'.format(root_path, lang, strings_name)
        # print(string_file_path)
        if os.path.exists(string_file_path):
            # 读取原有翻译文本
            content_dict = load_localize_content(string_file_path)
            # 添加excel中的翻译文本
            for j in range(2, max_row + 1):
                key = ws.cell(row=j, column=1).value
                if key is None:
                    continue
                value = ws.cell(row=j, column=i).value
                if value is None:
                    value = ""
                if value != "":
                    # print(value)
                    # print(len(value))
                    key_placeholder_list = re.findall(format_string_placeholder_rex, key, re.RegexFlag.DOTALL)
                    if isinstance(value, str):
                        value_placeholder_list = re.findall(format_string_placeholder_rex, value, re.RegexFlag.DOTALL)
                    else:
                        print("非字符串内容: {0} {1} {2}".format(j, lang, value))
                        value_placeholder_list = re.findall(format_string_placeholder_rex, "{0}".format(value),
                                                            re.RegexFlag.DOTALL)
                    if key_placeholder_list != value_placeholder_list:
                        print("占位符不匹配 模块: {0} 位置: 第{1}行 语言: {2}".format(mode_name, j, lang))
                        value = ""
                if (key == value or key_discard_placeholders.__contains__(value)) and not_wrap_key:
                    # 当不需要引号包裹时，翻译为空或key时代表这条不需要翻译
                    continue
                content_dict[key] = value
            # print(string_file_path)
            # 将翻译文本写入strings文件
            with open(string_file_path, 'wt') as string_file:
                sorted_keys = sorted(content_dict.keys())
                for key in sorted_keys:
                    value = content_dict[key]
                    if not not_wrap_key:
                        temp = '"{0}" = "{1}";\n'.format(key, value)
                    else:
                        temp = '{0} = "{1}";\n'.format(key, value)
                    string_file.write(temp)


def parse_argv():
    __parse = argparse.ArgumentParser()
    __parse.add_argument("modeName", help="翻译模块名称")
    __parse.add_argument("stringsName", help="翻译文件名称")
    __parse.add_argument("path", help="翻译文件路径")
    __parse.add_argument("-forImport", help="将翻译导入项目", action="store_true")
    __parse.add_argument("-sortByEmptyContent", help="是否根据未翻译内容排序", action="store_true")
    __parse.add_argument("-notWrapKey", help="不使用引号包裹key值", action="store_true")
    __parse.add_argument("-excelDir", help="翻译Excel文件目录", default="/Users/ldc/Desktop/翻译")
    return __parse.parse_args()


if __name__ == '__main__':
    argv = parse_argv()
    mode_name = argv.modeName
    home_dir = argv.excelDir
    excel_path = home_dir + "/翻译汇总.xlsx"
    strings_name = argv.stringsName
    root_path = subprocess.getoutput("cd {0};pwd".format(argv.path))
    if argv.forImport:
        localize_import(root_path, mode_name, strings_name, argv.notWrapKey)
    else:
        if not os.path.exists(home_dir):
            os.mkdir(home_dir)
        localize_export(root_path, mode_name, strings_name, argv.sortByEmptyContent)
